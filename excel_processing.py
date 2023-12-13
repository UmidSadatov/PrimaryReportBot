from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pyexcel as p


def convert_to_xlsx(filename: str):
    if filename.split('.')[-1] == 'xls':
        p.save_book_as(file_name=filename,
                       dest_file_name=filename + 'x')
        return filename + 'x'


class ExcelBook:
    def __init__(self, filename: str):
        extension = filename.split('.')[-1].lower()

        if extension not in ['xlsx', 'xls']:
            raise AttributeError('File extention must be .xlsx or .xls')
        elif extension == 'xls':
            self.filename = convert_to_xlsx(filename)
        else:
            self.filename = filename
        self.book = load_workbook(self.filename)

    def get_cell_value(self, cell: str, sheet_index=0, round_value=True):

        try:
            sheet = self.book.worksheets[sheet_index]
        except IndexError:
            return None

        value = sheet[cell].value if sheet[cell].value is not None else 0

        if type(value) == float and round_value:
            value = round(value, 2)

        return value

    def set_cell_value(self, cell, value, sheet_index=0):
        sheet = self.book.worksheets[sheet_index]
        sheet[cell].value = value

    def get_max_row(self, sheet_index=0):
        sheet = self.book.worksheets[sheet_index]
        return sheet.max_row


def make_report_excel(report_dict, sum_numbers, new_file_name):
    # print(report_dict)
    wb = Workbook()

    green_fill = PatternFill(start_color='ADFF2F',
                             end_color='ADFF2F',
                             fill_type='solid')

    yellow_fill = PatternFill(start_color='F8FF00',
                             end_color='F8FF00',
                             fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    bold_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )

    for sheet in report_dict:

        ws = wb.create_sheet(sheet)
        wb.active = ws

        ws.column_dimensions['A'].width = 40
        ws.merge_cells('A1:A2')
        ws['A1'], ws['A1'].font = 'Наименование', Font(bold=True)
        ws['A1'].alignment = \
            Alignment(horizontal='center', vertical='center')
        ws['A1'].border, ws['A2'].border = thin_border, thin_border

        ws.merge_cells('B1:C1')
        ws['B1'] = 'Остаток в начале'
        ws['B1'].font = Font(bold=True)
        ws['B1'].alignment = Alignment(horizontal='center')
        ws['B1'].border, ws['C1'].border = thin_border, thin_border

        ws['B2'], ws['C2'] = 'кол-во', 'стоимость'
        ws['B2'].border, ws['C2'].border = thin_border, thin_border
        ws['B2'].alignment = Alignment(horizontal='center')
        ws['C2'].alignment = Alignment(horizontal='center')
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        ws.merge_cells('D1:E1')
        ws['D1'] = 'Продажа'
        ws['D1'].font = Font(bold=True)
        ws['D1'].alignment = Alignment(horizontal='center')
        ws['D1'].border, ws['E1'].border = thin_border, thin_border

        ws['D2'], ws['E2'] = 'кол-во', 'стоимость'
        ws['D2'].border, ws['E2'].border = thin_border, thin_border
        ws['D2'].alignment = Alignment(horizontal='center')
        ws['E2'].alignment = Alignment(horizontal='center')
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        ws.merge_cells('F1:G1')
        ws['F1'] = 'Остаток в конце'
        ws['F1'].font = Font(bold=True)
        ws['F1'].alignment = Alignment(horizontal='center')
        ws['F1'].border, ws['G1'].border = thin_border, thin_border

        ws['F2'], ws['G2'] = 'кол-во', 'стоимость'
        ws['F2'].border, ws['G2'].border = thin_border, thin_border
        ws['F2'].alignment = Alignment(horizontal='center')
        ws['G2'].alignment = Alignment(horizontal='center')
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15

        row = 3

        balance_beginning_sum = 0
        balance_beginning_price_sum = 0

        consumption_sum = 0
        consumption_price_sum = 0

        balance_end_sum = 0
        balance_end_price_sum = 0

        for group in report_dict[sheet]:

            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = group
            ws[f'A{row}'].fill = yellow_fill

            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].border = thin_border
            ws[f'C{row}'].border = thin_border
            ws[f'D{row}'].border = thin_border
            ws[f'E{row}'].border = thin_border
            ws[f'F{row}'].border = thin_border
            ws[f'G{row}'].border = thin_border

            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].alignment = Alignment(horizontal='center')

            row += 1

            for name in report_dict[sheet][group]:
                ws[f'A{row}'] = name
                ws[f'A{row}'].border = thin_border

                ws[f'B{row}'] = \
                    report_dict[sheet][group][name]['balance_beginning']
                ws[f'B{row}'].border = thin_border
                ws[f'B{row}'].number_format = '# ###,00'
                balance_beginning_sum += \
                    report_dict[sheet][group][name]['balance_beginning']

                ws[f'C{row}'] = \
                    report_dict[sheet][group][name]['balance_beginning_price']
                ws[f'C{row}'].fill = green_fill
                ws[f'C{row}'].border = thin_border
                ws[f'C{row}'].number_format = '# ###,00'
                balance_beginning_price_sum += \
                    report_dict[sheet][group][name]['balance_beginning_price']

                ws[f'D{row}'] = report_dict[sheet][group][name]['consumption']
                ws[f'D{row}'].border = thin_border
                ws[f'D{row}'].number_format = '# ###,00'
                consumption_sum += \
                    report_dict[sheet][group][name]['consumption']

                ws[f'E{row}'] = \
                    report_dict[sheet][group][name]['consumption_price']
                ws[f'E{row}'].fill = green_fill
                ws[f'E{row}'].border = thin_border
                ws[f'E{row}'].number_format = '# ###,00'
                consumption_price_sum += \
                    report_dict[sheet][group][name]['consumption_price']

                ws[f'F{row}'] = \
                    report_dict[sheet][group][name]['balance_end']
                ws[f'F{row}'].border = thin_border
                ws[f'F{row}'].number_format = '# ###,00'
                balance_end_sum += \
                    report_dict[sheet][group][name]['balance_end']

                ws[f'G{row}'] = \
                    report_dict[sheet][group][name]['balance_end_price']
                ws[f'G{row}'].fill = green_fill
                ws[f'G{row}'].border = thin_border
                ws[f'G{row}'].number_format = '# ###,00'
                balance_end_price_sum += \
                    report_dict[sheet][group][name]['balance_end_price']

                row += 1

        ws[f'A{row}'] = 'ИТОГО'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = thin_border

        ws[f'B{row}'] = balance_beginning_sum
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].border = thin_border

        ws[f'C{row}'] = balance_beginning_price_sum
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'C{row}'].border = thin_border
        ws[f'C{row}'].fill = green_fill

        ws[f'D{row}'] = consumption_sum
        ws[f'D{row}'].font = Font(bold=True)
        ws[f'D{row}'].border = thin_border

        ws[f'E{row}'] = consumption_price_sum
        ws[f'E{row}'].font = Font(bold=True)
        ws[f'E{row}'].border = thin_border
        ws[f'E{row}'].fill = green_fill

        ws[f'F{row}'] = balance_end_sum
        ws[f'F{row}'].font = Font(bold=True)
        ws[f'F{row}'].border = thin_border

        ws[f'G{row}'] = balance_end_price_sum
        ws[f'G{row}'].font = Font(bold=True)
        ws[f'G{row}'].border = thin_border
        ws[f'G{row}'].fill = green_fill

    wb.remove(wb.worksheets[0])
    wb.save(new_file_name)
